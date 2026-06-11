"""
tests/test_study_export.py — §7.12 standalone study CSV / Excel exports.

Drives the same study_hooks() path the UI uses, builds a real study for
each kind (sensitivity / sweep / scenarios), and asserts:
  • CSV has self-describing metadata header + the expected table body.
  • Excel has a Metadata sheet + a Study sheet (chart + table).
  • The standalone Study sheet matches the one build_excel attaches.
  • Empty study raises a clear ValueError.
  • Disk persistence round-trips through pickle.
"""
import sys, os, json, pickle, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import datetime as _dt
import dataclasses as _dc
import pytest
import pandas as pd

import nexa_toolkit.engines                                  # noqa: registers
from nexa_toolkit.framework             import get
from nexa_toolkit.reporting.study_export import study_to_csv, study_to_xlsx
from nexa_toolkit.reporting.generic_report import build_excel

from nexablock.studies import (ParameterSweep, OneAtATimeSensitivity, ScenarioRunner)


# ── fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def engine():
    return get("gt_system_v2")


@pytest.fixture(scope="module")
def hooks(engine):
    return engine.study_hooks()


@pytest.fixture(scope="module")
def vals(engine):
    return engine.defaults()


def _enrich(study: dict, engine, hooks, vals) -> dict:
    """Add the metadata the UI normally stashes alongside the result."""
    return {**study,
            "engine_key":  engine.key,
            "engine_name": engine.name,
            "timestamp":   _dt.datetime.now().isoformat(timespec="seconds"),
            "base_params": _dc.asdict(hooks["make_params"](vals))}


def _sensitivity(engine, hooks, vals) -> dict:
    params = hooks["make_params"](vals)
    sens = OneAtATimeSensitivity(
        builder=hooks["builder"], base_params=params, kpi_fn=hooks["kpi_fn"],
        bounds=hooks["bounds"], step_override=hooks["step_override"],
    ).run(inputs=hooks["sensitivity_inputs"], kpis=hooks["kpis"])
    return _enrich({"kind": "sensitivity", "result": sens,
                    "chart_kwargs": {"kpi": hooks["kpis"][0]},
                    "label": f"Sensitivity tornado — {hooks['kpis'][0]}"},
                   engine, hooks, vals)


def _sweep(engine, hooks, vals) -> dict:
    params = hooks["make_params"](vals)
    lo, hi = hooks["bounds"]["load_pct"]
    values = [lo + (hi - lo) * i / 9 for i in range(10)]
    swp = ParameterSweep(hooks["builder"], params, hooks["kpi_fn"]).run(
        {"load_pct": values})
    return _enrich({"kind": "sweep", "result": swp,
                    "chart_kwargs": {"kpis": hooks["kpis"],
                                     "title": "Sweep over load_pct"},
                    "label": "Sweep over load_pct (10 points)"},
                   engine, hooks, vals)


def _scenarios(engine, hooks, vals) -> dict:
    params = hooks["make_params"](vals)
    res = ScenarioRunner(hooks["builder"], params, hooks["kpi_fn"]).run(
        hooks["scenarios"])
    return _enrich({"kind": "scenarios", "result": res,
                    "chart_kwargs": {"kpis": hooks["kpis"],
                                     "title": "Scenario comparison"},
                    "label": "Scenario comparison (ratio vs base)"},
                   engine, hooks, vals)


# ── CSV: metadata header + table body ────────────────────────────────────────

def test_csv_metadata_header_present(engine, hooks, vals, tmp_path):
    s = _sensitivity(engine, hooks, vals)
    p = tmp_path / "sens.csv"
    study_to_csv(s, str(p))
    text = p.read_text(encoding="utf-8")
    head = text.splitlines()[:7]
    assert any(l.startswith("# study: sensitivity")   for l in head)
    assert any(l.startswith("# engine: gt_system_v2") for l in head)
    assert any(l.startswith("# timestamp: ")          for l in head)
    assert any(l.startswith("# base_params: ")        for l in head)
    # base_params JSON parses and contains a known field
    bp_line = next(l for l in head if l.startswith("# base_params: "))
    bp = json.loads(bp_line[len("# base_params: "):])
    assert "load_pct" in bp


def test_csv_sensitivity_columns(engine, hooks, vals, tmp_path):
    s = _sensitivity(engine, hooks, vals)
    p = tmp_path / "sens.csv"
    study_to_csv(s, str(p))
    df = pd.read_csv(p, comment="#")
    for col in ("input", "kpi", "elasticity"):
        assert col in df.columns, f"missing {col!r}"
    assert len(df) >= len(hooks["sensitivity_inputs"]) * len(hooks["kpis"]) // 2


def test_csv_sweep_columns(engine, hooks, vals, tmp_path):
    s = _sweep(engine, hooks, vals)
    p = tmp_path / "swp.csv"
    study_to_csv(s, str(p))
    df = pd.read_csv(p, comment="#")
    assert any(c.startswith("in.")  for c in df.columns), df.columns.tolist()
    assert any(c.startswith("kpi.") for c in df.columns)
    assert len(df) == 10


def test_csv_scenarios_column(engine, hooks, vals, tmp_path):
    s = _scenarios(engine, hooks, vals)
    p = tmp_path / "scen.csv"
    study_to_csv(s, str(p))
    df = pd.read_csv(p, comment="#")
    assert "scenario" in df.columns
    assert set(df["scenario"]) == set(hooks["scenarios"].keys())


# ── XLSX: metadata + study sheets ────────────────────────────────────────────

def _open(path):
    from openpyxl import load_workbook
    return load_workbook(path)


def test_xlsx_has_metadata_and_study_sheets(engine, hooks, vals, tmp_path):
    s = _sweep(engine, hooks, vals)
    p = tmp_path / "swp.xlsx"
    study_to_xlsx(s, str(p))
    wb = _open(p)
    assert {"Metadata", "Study"}.issubset(set(wb.sheetnames))

    ws_m = wb["Metadata"]
    # row 3 onwards: key in col A, value in col B
    keys = [ws_m.cell(r, 1).value for r in range(3, 30) if ws_m.cell(r, 1).value]
    assert "kind" in keys
    assert "engine" in keys
    assert "timestamp" in keys
    assert any(k.startswith("base_params.") for k in keys), keys

    ws_s = wb["Study"]
    assert len(ws_s._images) >= 1, "no embedded chart in standalone Study sheet"
    assert ws_s.cell(24, 1).value == "Study data"
    headers = [ws_s.cell(25, j).value for j in range(1, 30) if ws_s.cell(25, j).value]
    assert any(h.startswith("in.")  for h in headers)
    assert any(h.startswith("kpi.") for h in headers)


def test_xlsx_study_sheet_matches_report_excel_study_sheet(engine, hooks,
                                                            vals, tmp_path):
    """Standalone and report-attached Study sheets must come from the same
    helper — same headers, same row count."""
    s = _sweep(engine, hooks, vals)
    standalone_p = tmp_path / "standalone.xlsx"
    study_to_xlsx(s, str(standalone_p))

    report_p = tmp_path / "report.xlsx"
    build_excel(engine, vals, engine.solve(vals), str(report_p), study=s)

    ws_a = _open(standalone_p)["Study"]
    ws_b = _open(report_p)["Study"]
    headers_a = [ws_a.cell(25, j).value for j in range(1, 30) if ws_a.cell(25, j).value]
    headers_b = [ws_b.cell(25, j).value for j in range(1, 30) if ws_b.cell(25, j).value]
    assert headers_a == headers_b, "headers diverge between standalone and report"

    def _count_data_rows(ws):
        n = 0
        while ws.cell(26 + n, 1).value is not None:
            n += 1
        return n
    assert _count_data_rows(ws_a) == _count_data_rows(ws_b)


# ── empty-case safety ────────────────────────────────────────────────────────

def test_empty_study_csv_raises(tmp_path):
    with pytest.raises(ValueError, match="empty study"):
        study_to_csv({}, str(tmp_path / "x.csv"))


def test_empty_study_xlsx_raises(tmp_path):
    with pytest.raises(ValueError, match="empty study"):
        study_to_xlsx({}, str(tmp_path / "x.xlsx"))


def test_missing_result_raises(tmp_path):
    """{"kind": "sweep"} but no result — still empty."""
    with pytest.raises(ValueError, match="empty study"):
        study_to_csv({"kind": "sweep"}, str(tmp_path / "x.csv"))


# ── disk persistence round-trip ──────────────────────────────────────────────

def test_pickle_roundtrip_preserves_kind_and_table(engine, hooks, vals, tmp_path):
    s = _sweep(engine, hooks, vals)
    blob = tmp_path / "s.pkl"
    with open(blob, "wb") as f:
        pickle.dump(s, f)
    with open(blob, "rb") as f:
        restored = pickle.load(f)
    assert restored["kind"]      == s["kind"]
    assert restored["timestamp"] == s["timestamp"]
    assert restored["base_params"]["load_pct"] == s["base_params"]["load_pct"]
    df_a = s["result"].as_dataframe()
    df_b = restored["result"].as_dataframe()
    assert df_a.shape == df_b.shape
    assert list(df_a.columns) == list(df_b.columns)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
