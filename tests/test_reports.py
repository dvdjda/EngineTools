"""
tests/test_reports.py — §7.11 reports with optional study attached.

Builds PDF and Excel reports through the v1 reporting layer with and
without a study attached, and asserts:
  • PDF without study: valid file (magic header + %%EOF trailer).
  • PDF with study: same validity, file is larger (the extra embedded
    study PNG adds bytes).
  • Excel without study: no "Study" sheet.
  • Excel with study: "Study" sheet exists, an image is embedded, and
    the underlying study table is present below the image with the
    expected columns.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest

import nexa_toolkit.engines             # noqa: F401  registers engines
from nexa_toolkit.framework            import get
from nexa_toolkit.reporting.generic_report import build_pdf, build_excel

from simulators.gt_system.system        import build_gt_system, summary
from nexablock.studies                   import (
    ParameterSweep, OneAtATimeSensitivity, ScenarioRunner)


# ── helpers ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def engine():
    return get("gt_system_v2")


@pytest.fixture(scope="module")
def hooks(engine):
    return engine.study_hooks()


@pytest.fixture(scope="module")
def vals(engine):
    return engine.defaults()


@pytest.fixture(scope="module")
def solved(engine, vals):
    return engine.solve(vals)


def _is_valid_pdf(path):
    with open(path, "rb") as f:
        head = f.read(5)
        f.seek(-1024, 2)         # last 1 KB for the EOF marker
        tail = f.read()
    return head.startswith(b"%PDF") and b"%%EOF" in tail


def _make_sensitivity_study(engine, hooks, vals):
    params = hooks["make_params"](vals)
    sens = OneAtATimeSensitivity(
        builder=hooks["builder"], base_params=params, kpi_fn=hooks["kpi_fn"],
        bounds=hooks["bounds"], step_override=hooks["step_override"],
    ).run(inputs=hooks["sensitivity_inputs"], kpis=hooks["kpis"])
    return {
        "kind":         "sensitivity",
        "result":       sens,
        "chart_kwargs": {"kpi": hooks["kpis"][0]},
        "label":        f"Sensitivity tornado — {hooks['kpis'][0]}",
    }


def _make_sweep_study(hooks, vals):
    params = hooks["make_params"](vals)
    lo, hi = hooks["bounds"]["load_pct"]
    values = [lo + (hi - lo) * i / 9 for i in range(10)]
    swp = ParameterSweep(hooks["builder"], params, hooks["kpi_fn"]).run(
        {"load_pct": values})
    return {
        "kind":         "sweep",
        "result":       swp,
        "chart_kwargs": {"kpis": hooks["kpis"], "title": "Sweep over load_pct"},
        "label":        "Sweep over load_pct (10 points)",
    }


def _make_scenarios_study(hooks, vals):
    params = hooks["make_params"](vals)
    res = ScenarioRunner(hooks["builder"], params, hooks["kpi_fn"]).run(
        hooks["scenarios"])
    return {
        "kind":         "scenarios",
        "result":       res,
        "chart_kwargs": {"kpis": hooks["kpis"], "title": "Scenario comparison"},
        "label":        "Scenario comparison (ratio vs base)",
    }


# ── PDF: with/without study ──────────────────────────────────────────────────

def test_pdf_without_study_is_valid(engine, vals, solved, tmp_path):
    p = tmp_path / "plain.pdf"
    build_pdf(engine, vals, solved, str(p), str(tmp_path / "c.png"))
    assert _is_valid_pdf(p)


def test_pdf_with_sensitivity_grows(engine, vals, solved, hooks, tmp_path):
    """The study chart is an extra embedded image; with-study should be larger."""
    p_plain = tmp_path / "plain.pdf"
    p_study = tmp_path / "study.pdf"
    build_pdf(engine, vals, solved, str(p_plain), str(tmp_path / "cp.png"))
    study = _make_sensitivity_study(engine, hooks, vals)
    build_pdf(engine, vals, solved, str(p_study), str(tmp_path / "cs.png"),
              study=study)
    assert _is_valid_pdf(p_study)
    size_plain = os.path.getsize(p_plain)
    size_study = os.path.getsize(p_study)
    assert size_study > size_plain + 8_000, (
        f"with-study PDF should be ≥ 8 KB larger; "
        f"plain={size_plain}, with study={size_study}")


@pytest.mark.parametrize("kind_fn", [
    pytest.param(_make_sensitivity_study, id="sensitivity"),
    pytest.param(lambda e, h, v: _make_sweep_study(h, v),   id="sweep"),
    pytest.param(lambda e, h, v: _make_scenarios_study(h, v), id="scenarios"),
])
def test_pdf_with_each_study_kind_is_valid(engine, vals, solved, hooks,
                                            tmp_path, kind_fn):
    study = kind_fn(engine, hooks, vals)
    p = tmp_path / f"pdf_{study['kind']}.pdf"
    build_pdf(engine, vals, solved, str(p), str(tmp_path / "c.png"),
              study=study)
    assert _is_valid_pdf(p)


# ── Excel: with/without study ────────────────────────────────────────────────

def _open_xlsx(path):
    from openpyxl import load_workbook
    return load_workbook(path)


def test_excel_without_study_has_no_study_sheet(engine, vals, solved, tmp_path):
    p = tmp_path / "plain.xlsx"
    build_excel(engine, vals, solved, str(p))
    wb = _open_xlsx(p)
    assert "Study" not in wb.sheetnames


def test_excel_with_sweep_has_sheet_chart_and_table(engine, vals, solved,
                                                    hooks, tmp_path):
    study = _make_sweep_study(hooks, vals)
    p = tmp_path / "with_sweep.xlsx"
    build_excel(engine, vals, solved, str(p), study=study)

    wb = _open_xlsx(p)
    assert "Study" in wb.sheetnames
    ws = wb["Study"]

    # Header cell with the study label
    assert ws["A1"].value and "Sweep" in ws["A1"].value

    # An image is embedded (the rendered sweep chart PNG)
    # openpyxl stashes images on ws._images
    assert len(ws._images) >= 1, "no embedded image in Study sheet"

    # The table lands at row 24 with headers in row 25
    assert ws.cell(24, 1).value == "Study data"
    headers = [ws.cell(25, j).value for j in range(1, 30) if ws.cell(25, j).value]
    assert any(h.startswith("in.")  for h in headers), f"no input cols: {headers}"
    assert any(h.startswith("kpi.") for h in headers), f"no KPI cols: {headers}"
    # At least one data row below — pick first KPI column and check a numeric
    in_col = next(j for j, h in enumerate(headers, 1) if h.startswith("in."))
    kpi_col = next(j for j, h in enumerate(headers, 1) if h.startswith("kpi."))
    first_in  = ws.cell(26, in_col).value
    first_kpi = ws.cell(26, kpi_col).value
    assert isinstance(first_in,  (int, float)) and first_in is not None
    assert isinstance(first_kpi, (int, float)) and first_kpi is not None


# ── Convergence-warning surfacing in reports ────────────────────────────────

def _build_failing_engine():
    """Tiny Engine that wraps the test_core recycle loop and forces non-
    convergence via a tol below machine epsilon. Used to check the
    convergence-warning path of build_excel end to end."""
    from nexa_toolkit.framework.contract import Engine, InputSpec, OutputSpec
    from nexablock.core.system   import System
    from nexablock.core.stream   import Stream, StreamKind
    from nexablock.core.recycle  import Recycle
    from tests.test_core         import Heater, Cooler

    class _FailEngine(Engine):
        key = "_test_fail_engine"
        name = "Failing recycle engine (test)"
        status = "draft"
        inputs = [InputSpec("dummy", "Dummy", "-", 0.0, -1, 1)]
        def solve(self, v):
            sys = System("forced-fail")
            heater  = sys.add(Heater(Q_kW=200.0))   # imbalanced duties →
            cooler  = sys.add(Cooler(Q_kW=50.0))    # no fixed point → diverges
            recycle = sys.add(Recycle(StreamKind.WATER_STEAM, tol=1e-4))
            seed = Stream.water_steam(mdot=2.0, T=300.0, P=2e5, h=1.2e5)
            recycle.inlets["inlet"].stream = seed
            sys.connect(heater.outlets["outlet"],  cooler.inlets["inlet"])
            sys.connect(cooler.outlets["outlet"],  recycle.inlets["inlet"])
            sys.connect(recycle.outlets["outlet"], heater.inlets["inlet"])
            solved = sys.solve()                    # default tol/max_iter
            return {"solved": solved, "kpis": {}}
        def outputs(self, r):
            return [OutputSpec("Dummy KPI", 1.0, "-", "verified", "{:.2f}")]

    return _FailEngine()


def test_xlsx_not_converged_shows_warning_cell(tmp_path):
    """Build an Excel for a non-converged solve; assert the NOT CONVERGED
    warning text and the convergence summary land as cell values, AND that
    the (single) result row's Basis cell reads 'unverified' (KPI flagged)."""
    from openpyxl import load_workbook
    e = _build_failing_engine()
    v = e.defaults()
    r = e.solve(v)
    p = tmp_path / "bad.xlsx"
    build_excel(e, v, r, str(p))

    ws = load_workbook(p).active
    cells = []
    for row in ws.iter_rows(values_only=True):
        cells.extend(str(c) for c in row if c is not None)
    joined = "\n".join(cells)
    assert "NOT CONVERGED" in joined, "warning text missing from Excel"
    assert "Convergence" in joined,    "Convergence band missing from Excel"
    assert "unverified"  in joined,    "result row Basis should be 'unverified'"


def test_xlsx_deficit_shows_warning_and_flags_kpis(engine, hooks, tmp_path):
    """gpu_it_kW=10000 forces a ~2600 kW deficit. Excel must carry the
    POWER DEFICIT warning, the Power balance band, the assumption line,
    and every Basis cell flipped to 'unverified' — even though the solve
    converges (feasibility independently flags KPIs)."""
    from openpyxl import load_workbook
    v = engine.defaults(); v["gpu_it_kW"] = 10000.0
    r = engine.solve(v)
    assert r["solved"].convergence.converged       # convergence still OK
    assert not r["feasibility"].feasible           # feasibility flips

    p = tmp_path / "deficit.xlsx"
    build_excel(engine, v, r, str(p))
    ws = load_workbook(p).active
    cells = []
    for row in ws.iter_rows(values_only=True):
        cells.extend(str(c) for c in row if c is not None)
    joined = "\n".join(cells)
    assert "POWER DEFICIT"   in joined, "deficit warning missing from Excel"
    assert "Power balance"   in joined, "Power balance band missing"
    assert "Assumption"      in joined, "assumption line missing"
    assert "GT-powered"      in joined, "assumption text missing"
    assert "unverified"      in joined, "KPI Basis cells should be flagged"
    # Every itemised aux line must appear in the breakdown.
    for line in ("LiBr pump electrical",
                 "Cooling tower fan electrical",
                 "GT auxiliaries",
                 "Plant BoP"):
        assert line in joined, f"breakdown line {line!r} missing from Excel"
    # Convergence still reads green — two separate truths.
    assert "no recycle loops" in joined


def test_xlsx_feasible_default_no_deficit_warning(engine, vals, solved, tmp_path):
    """Defaults are feasible; the Power balance band exists but no
    POWER DEFICIT warning."""
    from openpyxl import load_workbook
    p = tmp_path / "feasible.xlsx"
    build_excel(engine, vals, solved, str(p))
    ws = load_workbook(p).active
    cells = []
    for row in ws.iter_rows(values_only=True):
        cells.extend(str(c) for c in row if c is not None)
    joined = "\n".join(cells)
    assert "Power balance" in joined
    assert "POWER DEFICIT" not in joined


def test_xlsx_converged_no_warning(engine, vals, solved, tmp_path):
    """Sanity counter-example: GT v2 acyclic solve produces a green convergence
    summary, no warning text in the workbook."""
    from openpyxl import load_workbook
    p = tmp_path / "good.xlsx"
    build_excel(engine, vals, solved, str(p))
    ws = load_workbook(p).active
    cells = []
    for row in ws.iter_rows(values_only=True):
        cells.extend(str(c) for c in row if c is not None)
    joined = "\n".join(cells)
    assert "no recycle loops" in joined
    assert "NOT CONVERGED" not in joined


def test_excel_with_scenarios_table_has_scenario_column(engine, vals, solved,
                                                        hooks, tmp_path):
    """ScenarioResult.as_dataframe is index=scenario name; we reset_index so
    the scenario name lands as a real column called 'scenario'."""
    study = _make_scenarios_study(hooks, vals)
    p = tmp_path / "with_scenarios.xlsx"
    build_excel(engine, vals, solved, str(p), study=study)
    wb = _open_xlsx(p)
    ws = wb["Study"]
    headers = [ws.cell(25, j).value for j in range(1, 30) if ws.cell(25, j).value]
    assert "scenario" in headers, f"no scenario column: {headers}"
    scen_col = headers.index("scenario") + 1
    names = {ws.cell(26 + i, scen_col).value for i in range(len(hooks["scenarios"]))}
    assert names == set(hooks["scenarios"].keys())


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
