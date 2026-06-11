"""
nexa_toolkit.reporting.study_export — standalone study downloads.

Two exporters that serialise the underlying table of a stored study
result, with light self-describing metadata (kind, engine, timestamp,
base params).

The Excel exporter reuses write_study_sheet from generic_report so a
standalone study workbook has the same "Study" sheet (chart + table)
as the one attached to a full PDF/Excel report — one source of truth.

    study = _LATEST_STUDY[engine_key]   # populated by the UI study buttons
    study_to_csv(study, "/tmp/my_study.csv")
    study_to_xlsx(study, "/tmp/my_study.xlsx")
"""
from __future__ import annotations
import json

from .generic_report import write_study_sheet


def _check_study(study) -> None:
    if not study or "result" not in study:
        raise ValueError("empty study: nothing to export")


def _metadata_pairs(study) -> list:
    """List of (label, value) rows for the metadata block."""
    base_params = study.get("base_params") or {}
    rows = [
        ("kind",        study.get("kind", "")),
        ("engine",      study.get("engine_key", "")),
        ("engine_name", study.get("engine_name", "")),
        ("label",       study.get("label", "")),
        ("timestamp",   study.get("timestamp", "")),
    ]
    for k, v in base_params.items():
        rows.append((f"base_params.{k}", v))
    return rows


# ── CSV ──────────────────────────────────────────────────────────────────────

def study_to_csv(study: dict, path: str) -> str:
    """Write study to CSV with `# `-prefixed metadata header lines."""
    _check_study(study)
    df = study["result"].as_dataframe()
    if getattr(df.index, "name", None) or df.index.dtype == object:
        df = df.reset_index().rename(columns={"index": "scenario"})

    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write(f"# study: {study.get('kind','')}\n")
        f.write(f"# engine: {study.get('engine_key','')}\n")
        f.write(f"# engine_name: {study.get('engine_name','')}\n")
        f.write(f"# label: {study.get('label','')}\n")
        f.write(f"# timestamp: {study.get('timestamp','')}\n")
        f.write(f"# base_params: {json.dumps(study.get('base_params', {}), default=str)}\n")
        f.write("#\n")
        df.to_csv(f, index=False)
    return path


# ── Excel ────────────────────────────────────────────────────────────────────

def study_to_xlsx(study: dict, path: str) -> str:
    """Workbook with two sheets:
       Metadata — key/value rows (kind, engine, timestamp, base_params)
       Study    — chart + table via write_study_sheet (same as build_excel)."""
    _check_study(study)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws_m = wb.active
    ws_m.title = "Metadata"
    ws_m["A1"] = "Study metadata"
    ws_m["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws_m["A1"].fill = PatternFill("solid", fgColor="2E4E7E")
    ws_m["A1"].alignment = Alignment(vertical="center", indent=1)
    ws_m.column_dimensions["A"].width = 30
    ws_m.column_dimensions["B"].width = 60

    for i, (k, v) in enumerate(_metadata_pairs(study), start=3):
        ws_m.cell(i, 1, k).font = Font(bold=True)
        ws_m.cell(i, 2, v if not isinstance(v, (list, dict)) else json.dumps(v, default=str))

    # Study sheet — reuses the helper that build_excel also calls.
    write_study_sheet(wb, study)

    wb.save(path)
    return path
