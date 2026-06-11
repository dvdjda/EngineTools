"""Excel report for a LiBr chiller sizing case."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

NAVY = "2E4E7E"
LIGHT = "EAF0F8"
WHITE = "FFFFFF"
GREY = "595959"
BLUE_TX = "0000FF"

thin = Side(style="thin", color="C9D2E0")
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def _band(ws, cell, text, span):
    ws[cell] = text
    ws[cell].font = Font(name="Arial", bold=True, size=12, color=WHITE)
    ws[cell].fill = PatternFill("solid", fgColor=NAVY)
    ws[cell].alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(f"{cell}:{span}")


def _hdr(ws, row):
    for col, t in zip("ABC", ("Quantity", "Value", "Unit")):
        c = ws[f"{col}{row}"]
        c.value = t
        c.font = Font(name="Arial", bold=True, color=NAVY)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.border = box
        c.alignment = Alignment(horizontal="left" if col == "A" else "center")


def build_excel(dp, r, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sizing"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    ws["A1"] = "Single-effect LiBr-H2O absorption chiller"
    ws["A1"].font = Font(name="Arial", bold=True, size=15, color=NAVY)
    ws["A2"] = "Screening sizing report  -  verify duties vs ChemCAD"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color=GREY)

    # design point
    _band(ws, "A4", "Design point (inputs)", "C4")
    _hdr(ws, 5)
    design = [
        ("Chilled-water supply", dp.t_chw_out_c, "degC"),
        ("Cooling-water inlet", dp.t_cw_in_c, "degC"),
        ("Heat source", dp.t_hot_c, "degC"),
        ("Cooling duty", dp.q_evap_kw, "kW"),
    ]
    row = 6
    for name, val, unit in design:
        ws[f"A{row}"] = name
        ws[f"B{row}"] = val
        ws[f"B{row}"].font = Font(name="Arial", color=BLUE_TX)  # editable input
        ws[f"C{row}"] = unit
        for col in "ABC":
            ws[f"{col}{row}"].border = box
            if col != "A":
                ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
        row += 1

    # results (engine outputs as values; derived metrics as formulas)
    _band(ws, "A11", "Results", "C11")
    _hdr(ws, 12)
    # primary engine outputs
    outputs = [
        ("Weak (dilute) solution", r["x_weak_pct"], "% LiBr"),
        ("Strong (conc.) solution", r["x_strong_pct"], "% LiBr"),
        ("Circulation ratio", r["circulation_ratio"], "-"),
        ("Evaporator duty", r["q_evap_kw"], "kW"),
        ("Condenser duty", r["q_cond_kw"], "kW"),
        ("Generator duty", r["q_gen_kw"], "kW"),
        ("Absorber duty", r["q_abs_kw"], "kW"),
        ("Crystallisation line", r["x_crystallisation_pct"], "% LiBr"),
    ]
    row = 13
    rowref = {}
    for name, val, unit in outputs:
        ws[f"A{row}"] = name
        ws[f"B{row}"] = round(val, 3)
        ws[f"C{row}"] = unit
        rowref[name] = row
        for col in "ABC":
            ws[f"{col}{row}"].border = box
            if col != "A":
                ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
        row += 1

    # derived metrics via formulas
    derived = [
        ("Concentration swing",
         f"=B{rowref['Strong (conc.) solution']}-B{rowref['Weak (dilute) solution']}", "%"),
        ("COP",
         f"=B{rowref['Evaporator duty']}/B{rowref['Generator duty']}", "-"),
        ("Energy balance check",
         f"=B{rowref['Evaporator duty']}+B{rowref['Generator duty']}"
         f"-B{rowref['Condenser duty']}-B{rowref['Absorber duty']}", "kW"),
        ("Crystallisation margin",
         f"=B{rowref['Crystallisation line']}-B{rowref['Strong (conc.) solution']}", "%"),
    ]
    for name, formula, unit in derived:
        ws[f"A{row}"] = name
        ws[f"B{row}"] = formula
        ws[f"C{row}"] = unit
        ws[f"A{row}"].font = Font(name="Arial", bold=True, color=NAVY)
        ws[f"B{row}"].font = Font(name="Arial", bold=True)
        for col in "ABC":
            ws[f"{col}{row}"].border = box
            if col != "A":
                ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
        row += 1

    # number formats
    for rr in range(6, row):
        ws[f"B{rr}"].number_format = "0.0##"

    # native chart of the four duties
    chart = BarChart()
    chart.type = "col"
    chart.title = "Component duties (kW)"
    chart.legend = None
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=2,
                     min_row=rowref["Evaporator duty"], max_row=rowref["Absorber duty"])
    cats = Reference(ws, min_col=1,
                     min_row=rowref["Evaporator duty"], max_row=rowref["Absorber duty"])
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, "E4")

    # default font on all used cells
    for r_ in ws.iter_rows():
        for c in r_:
            if c.font and c.font.name != "Arial":
                f = c.font
                c.font = Font(name="Arial", bold=f.bold, italic=f.italic,
                              size=f.size, color=f.color)
    wb.save(path)
    return path
